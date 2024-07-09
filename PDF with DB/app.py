import gradio as gr
# from gradio_calendar import Calendar
from datetime import datetime, timedelta
# from sqlalchemy import create_engine, select, distinct, or_
# from sqlalchemy.orm import sessionmaker
from typing import List
import os, docx, zipfile
from pdfrw import PdfReader, PdfWriter
# from common.models import Company, Entity, EntityType, Role, RoleType, Stockholder
# from common.fetch_all_from_db import main as retrieve_all
# from common.save_to_db import save_company_data, save_entity_data, save_role_data, save_stockholder_data
# from scripts.create.action_of_incorporator import main as create_action_of_incorporator
# from scripts.create.board_consent_adopt_stock_plan import main as create_board_consent_adopt_stock_plan
# from scripts.create.board_consent_bank_authorization import main as create_board_consent_bank_authorization
# from scripts.create.board_consent_formation import main as create_board_consent_formation
# from scripts.create.ca_certificate_of_incorporation import main as create_ca_certificate_of_incorporation
# from scripts.create.ca_corporate_bylaws import main as create_ca_corporate_bylaws
# from scripts.create.ca_indemnification_agreement import main as create_ca_indemnification_agreement
# from scripts.create.de_certificate_of_incorporation_par_value_stocks import main as create_de_certificate_of_incorporation
# from scripts.create.de_corporate_bylaws import main as create_de_corporate_bylaws
# from scripts.create.de_indemnification_agreement import main as create_de_indemnification_agreement
# from scripts.create.ein_application_authorization import main as create_ein_application_authorization
# from scripts.create.entity_stock_purchase_agreement import main as create_entity_stock_purchase_agreement
# from scripts.create.ip_assignment_agreement import main as create_ip_assignment_agreement
# from scripts.create.natural_person_no_vesting_stock_purchase_agreement import main as create_natural_person_no_vesting_stock_purchase_agreement
# from scripts.create.natural_person_vesting_stock_purchase_agreement import main as create_natural_person_vesting_stock_purchase_agreement
# from scripts.create.secretary_certificate_adopt_bylaws import main as create_secretary_certificate_adopt_bylaws
from scripts.ss4 import main as create_ss4
# from scripts.create.stock_certificate import main as create_stock_certificate
# from scripts.create.stock_incentive_plan import main as create_stock_incentive_plan
# from scripts.create.stockholder_consent_adopt_stock_plan import main as create_stockholder_consent_adopt_stock_plan
# from scripts.create.stockholder_consent_indemnification_agreement import main as create_stockholder_consent_indemnification_agreement
from openpyxl import load_workbook, Workbook
import mysql.connector

states = [
    'Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado',
    'Connecticut', 'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho',
    'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana',
    'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota',
    'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada',
    'New Hampshire', 'New Jersey', 'New Mexico', 'New York',
    'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon',
    'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota',
    'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington',
    'West Virginia', 'Wisconsin', 'Wyoming'
]

forms = [
    "CERTIFICATE OF INCORPORATION",
    "CORPORATE BYLAWS",
    "CONSENT OF INCORPORATOR",
    "INITIAL ORGANIZATIONAL BOARD CONSENT",
    "BANK CONSENT",
    "OFFICER AND DIRECTOR INDEMNIFICATION AGREEMENT",
    "STOCK PURCHASE AGREEMENT",
    "IP CONTRIBUTION AND ASSIGNMENT AGREEMENT",
    "EMPLOYER IDENTIFICATION NUMBER (EIN) AUTHORIZATION",
    "STOCK CERTIFICATE",
    "STOCKHOLDER INDEMNIFICATION AGREEMENT",
    "STOCK INCENTIVE PLAN",
    "BOARD CONSENT ADOPT STOCK PLAN",
    "STOCKHOLDER CONSENT ADOPT STOCK PLAN",
    "SECRETARY CERTIFICATE ADOPT BYLAWS"
]

TEMPLATE_DIR = "templates"
DE_COI_FILENAME = "DE - Certificate of Incorporation - Par Value Stock - Forms"
DE_CORPORATE_BYLAWS_FILENAME = "DE - Corporate Bylaws - Right of First Refusal - Forms"
CA_COI_FILENAME = "CA - Certificate of Incorporation"
CA_CORPORATE_BYLAWS_FILENAME = "CA - Corporate Bylaws"
CONSENT_OF_INCORPORATOR_FILENAME = "Action of Incorporator.1"
INITIAL_ORGANIZATIONAL_BOARD_CONSENT = "Board Consent - Formation.1"

de_coi_path = os.path.join(TEMPLATE_DIR, "DE - Certificate of Incorporation - Par Value Stock - Forms.docx")
de_bylaws_path = os.path.join(TEMPLATE_DIR, "DE - Corporate Bylaws - Right of First Refusal - Forms.docx")
ca_coi_path = os.path.join(TEMPLATE_DIR, "CA - Certificate of Incorporation.pdf")
ca_bylaws_path = os.path.join(TEMPLATE_DIR, "CA - Corporate Bylaws.1.docx")
action_of_incorporator_path = os.path.join(TEMPLATE_DIR, "Action of Incorporator.1.docx")
board_consent_formation_path = os.path.join(TEMPLATE_DIR, "Board Consent - Formation.1.docx")

board_consent_bank_authorization_path = os.path.join(TEMPLATE_DIR, "Board Consent - Bank Authorization.1.docx")
ca_indemnification_agreement_path = os.path.join(TEMPLATE_DIR, "CA - Officer and Director Indemnification Agreement.1.docx")
de_indemnification_agreement_path = os.path.join(TEMPLATE_DIR, "DE - Officer and Director Indemnification Agreement.1.docx")
stock_purchase_agreement_entity_path = os.path.join(TEMPLATE_DIR, "Entity Stock Purchase Agreement.1.docx")
stock_purchase_agreement_person_vest_path = os.path.join(TEMPLATE_DIR, "Natural Person - Stock Purchase Agreement - Vesting and Right of First Refusal.1.docx")
stock_purchase_agreement_person_no_vest_path = os.path.join(TEMPLATE_DIR, "Natural Person (No vesting, with Right of First Refusal) - Stock Purchase Agreement.1.docx")
ip_assignment_agreement_path = os.path.join(TEMPLATE_DIR, "IP Assignment Agreement (goes with Founder Stock Purchase Agreement).1.docx")
ein_path = os.path.join(TEMPLATE_DIR, "EIN Application Authorization.1.docx")
stock_certificate_path = os.path.join(TEMPLATE_DIR, "Stock Certificate.docx")
stockholder_indemnification_agreement_path = os.path.join(TEMPLATE_DIR, "Stockholder Consent - Indemnification Agreement.1.docx")
stock_incentive_plan_path = os.path.join(TEMPLATE_DIR, "Stock Incentive Plan.1.docx")
board_consent_adopt_stock_plan_path = os.path.join(TEMPLATE_DIR, "Board Consent - Adopt Stock Plan.1.docx")
stockholder_consent_adopt_stock_plan_path = os.path.join(TEMPLATE_DIR, "Stockholder Consent - Adopt Stock Plan.1.docx")
secretary_certificate_adopt_bylaws_path = os.path.join(TEMPLATE_DIR, "Secretary Certificate - Adopt Bylaws.1.docx")
ss4_path = os.path.join(TEMPLATE_DIR, "SS4.pdf")
invoice_path = os.path.join(TEMPLATE_DIR, "Invoice.xlsx")

# engine = create_engine('mysql+pymysql://dbuser:dbuser@localhost/test2', pool_recycle=3600)
# Session = sessionmaker(bind=engine)

VISIBLE = False
MAX_NUM_ENTITIES = 10
MAX_NUM_OTHER_OFFICERS = 5


# def update_document(company_id: int, download_checkbox: List[str]):
#     with Session() as session:
#         company = session.scalars(select(Company).where(Company.company_id == company_id)).one()

#         founders = session.execute(
#             select(
#                 Entity.name, 
#                 Entity.address, 
#                 Stockholder.number_of_shares, 
#                 Stockholder.purchase_price, 
#                 Stockholder.aggregate_total_price
#             )
#             .join(Stockholder).join(Role).where(
#                 Role.company_id == company_id,
#                 Role.role_type == RoleType.founder,
#             )).all()

#         directors = session.scalars(
#             select(Entity.name).join(Role).where(
#                 Role.role_type == RoleType.director, 
#                 Role.company_id == company_id
#             )).all()

#         chairman_name = session.scalars(
#             select(Entity.name).join(Role).where(
#                 Role.title == "Chairman", 
#                 Role.company_id == company_id
#             )).first()
#         ceo_name = session.scalars(
#             select(Entity.name).join(Role).where(
#                 Role.title == "CEO", 
#                 Role.company_id == company_id
#             )).first()
#         cfo_name = session.scalars(
#             select(Entity.name).join(Role).where(
#                 Role.title == "CFO", 
#                 Role.company_id == company_id
#             )).first()
#         secretary_name = session.scalars(
#             select(Entity.name).join(Role).where(
#                 Role.title == "Secretary", 
#                 Role.company_id == company_id
#             )).first()
#         account_open_officers = session.execute(
#             select(Entity.name, Role.title).join(Role).where(
#                 Role.permissions.ilike("%account_open%"),
#                 Role.company_id == company_id
#             )).all()
#         account_draw_officers = session.execute(
#             select(Entity.name, Role.title).join(Role).where(
#                 Role.permissions.ilike("%account_draw%"),
#                 Role.company_id == company_id
#             )).all()
        
#         excluded_titles = ["CEO", "CFO", "Secretary"]
#         other_officers = session.execute(
#             select(
#                 Entity.name, 
#                 Role.title
#             ).join(Role).where(
#                 Role.role_type == RoleType.officer, 
#                 Role.title.not_in(excluded_titles), 
#                 Role.company_id == company_id
#             )).all()
        
#         stockholders = session.execute(
#             select(
#                 Entity.name, 
#                 Entity.entity_type, 
#                 Entity.formation_state, 
#                 Entity.address, 
#                 Entity.spouse_name,
#                 Stockholder.number_of_shares, 
#                 Stockholder.purchase_price,
#                 Stockholder.aggregate_total_price,
#                 Stockholder.issue_date,
#                 Stockholder.total_vesting_at_cliff,
#                 Stockholder.cliff_anniversary,
#                 Stockholder.monthly_vesting_after_cliff
#             ).join(Entity).where(
#                 Stockholder.company_id == company_id
#             )).all()

#         indemnitees = session.execute(
#             select(
#                 distinct(Entity.name), 
#                 Entity.address
#             ).join(Role).where(
#                 Role.company_id == company_id,
#                 or_(
#                     Role.role_type == RoleType.director, 
#                     Role.role_type == RoleType.officer
#                 )
#             )).all()

#     ### Create Company Directory
#     if not os.path.exists(f"outputs/{company.company_name}"):
#         os.mkdir(f"outputs/{company.company_name}")
    
#     ### Create Action Of Incorporator
#     if "CONSENT OF INCORPORATOR" in download_checkbox:
#         conofinc_doc = docx.Document(action_of_incorporator_path)
#         values = {
#             "formation_state": company.formation_state,
#             "company_name": company.company_name,
#             "incorporator_name": company.incorporator_name,
#             "formation_date": company.formation_date.strftime("%B %d, %Y"),
#             "num_directors": len(directors)
#         }
#         updated_doc = create_action_of_incorporator(doc=conofinc_doc, values=values, directors=directors)
#         output_path = f"outputs/{company.company_name}/{CONSENT_OF_INCORPORATOR_FILENAME}.docx"
#         updated_doc.save(output_path)

#     ### Create Board Consent Formation
#     if "INITIAL ORGANIZATIONAL BOARD CONSENT" in download_checkbox:
#         bcform_doc = docx.Document(board_consent_formation_path)
#         values = {
#             "formation_state": company.formation_state,
#             "company_name": company.company_name,
#             "par_value": f"{company.par_value:.5f}",
#             "formation_date": company.formation_date.strftime("%B %d, %Y"),
#             "num_directors": len(directors),
#             "purchase_price": f"{stockholders[0].purchase_price:.5f}",
#             "company_address": company.company_address,
#             "fy_end": company.fiscal_year_end
#         }
#         officer_mapping = {
#             "Chairman of the Board": chairman_name,
#             "President and Chief Executive Officer": ceo_name,
#             "Treasurer and Chief Financial Officer": cfo_name,
#             "Corporate Secretary": secretary_name
#         }
#         additional_officers = []
#         for name, post in other_officers:
#             additional_officers.append({
#                 "post": post,
#                 "name": name
#             })
#         stockholder_rows = []
#         for name, _, _, _, _, num_shares, _, price, _, _, _, _ in stockholders:
#             stockholder_rows.append({
#                 "purchaser_name": name,
#                 "purchaser_share_purchased": str(num_shares),
#                 "purchaser_total_price": str(price),
#                 "form_of_agreement": "Attached as Exhibit Form of Stock Purchase Agreement"
#             })
#         updated_doc = create_board_consent_formation(doc=bcform_doc, values=values, officer_mapping=officer_mapping, additional_officers=additional_officers, stockholder_rows=stockholder_rows, directors=directors)
#         output_path = f"outputs/{company.company_name}/{INITIAL_ORGANIZATIONAL_BOARD_CONSENT}.docx"
#         updated_doc.save(output_path)

#     ### Create Certificate Of Incorporation
#     if "CERTIFICATE OF INCORPORATION" in download_checkbox:
#         if company.formation_state == "Delaware":
#             coi_doc = docx.Document(de_coi_path)
#             values = {
#                 "company_name": company.company_name,
#                 "agent_address": company.agent_address,
#                 "agent_city": company.agent_city,
#                 "agent_county": company.agent_county,
#                 "agent_zipcode": company.agent_zipcode,
#                 "agent_name": company.agent_name,
#                 "num_common_stock": company.num_common_stock,
#                 "par_value": f"{company.par_value:.5f}",
#                 "incorporator_name": company.incorporator_name,
#                 "incorporator_address": company.incorporator_address,
#                 "submission_date": company.submission_date.strftime("%B %d, %Y"),
#             }
#             updated_doc = create_de_certificate_of_incorporation(doc=coi_doc, values=values)
#             output_path = f"outputs/{company.company_name}/{DE_COI_FILENAME}.docx"
#             updated_doc.save(output_path)
#         if company.formation_state == "California":
#             coi_pdf = PdfReader(ca_coi_path)
#             values = {
#                 "(1nameLine1)": company.company_name,
#                 "(2aStreetAddress)": company.company_address,
#                 "(3aAgentsName)": company.agent_name,
#                 "(3bAgentsAddress)": company.agent_address,
#                 "(3bCity)": company.agent_city,
#                 "(3bZip)": company.agent_zipcode,
#                 "(4Shares)": company.num_common_stock,
#                 "(6TypeNameOfSigner)": company.incorporator_name
#             }
#             updated_pdf = create_ca_certificate_of_incorporation(pdf=coi_pdf, values=values)
#             output_path = f"outputs/{company.company_name}/{CA_COI_FILENAME}.pdf"
#             PdfWriter().write(output_path, updated_pdf)

    ### Create Corporate Bylaws
    # if "CORPORATE BYLAWS" in download_checkbox:
    #     if company.formation_state == "Delaware":
    #         bylaws_doc = docx.Document(de_bylaws_path)
    #         values = {
    #             "company_name": company.company_name,
    #             "agent_address": company.agent_address,
    #         }
    #         updated_doc = create_de_corporate_bylaws(doc=bylaws_doc, values=values)
    #         output_path = f"outputs/{company.company_name}/{DE_CORPORATE_BYLAWS_FILENAME}.docx"
    #         updated_doc.save(output_path) 
    #     if company.formation_state == "California":
    #         bylaws_doc = docx.Document(ca_bylaws_path)
    #         values = {
    #             "company_name": company.company_name,
    #             "num_directors": len(directors),
    #         }
    #         updated_doc = create_ca_corporate_bylaws(doc=bylaws_doc, values=values)
    #         output_path = f"outputs/{company.company_name}/{CA_CORPORATE_BYLAWS_FILENAME}.docx"
    #         updated_doc.save(output_path) 
    
#     ### Create Bank Consent
#     if "BANK CONSENT" in download_checkbox:
#         board_consent_bank_authorization_doc = docx.Document(board_consent_bank_authorization_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "company_bank": company.company_bank,
#             "formation_date": company.formation_date.strftime("%B %d, %Y")
#         }
#         if account_open_officers:
#             account_open_officers = [{"name": name, "post": title} for (name, title) in account_open_officers]
#         if account_draw_officers:
#             account_draw_officers = [{"name": name, "post": title} for (name, title) in account_draw_officers]
#         updated_doc = create_board_consent_bank_authorization(doc=board_consent_bank_authorization_doc, values=values, account_open_officers=account_open_officers, account_draw_officers=account_draw_officers, directors=directors)
#         output_path = f"outputs/{company.company_name}/Bank Consent.docx"
#         updated_doc.save(output_path)

#     ### Create Officer and Director Indemnification Agreements
#     if "OFFICER AND DIRECTOR INDEMNIFICATION AGREEMENT" in download_checkbox:
#         if not os.path.exists(f"outputs/{company.company_name}/Officer and Director Indemnification Agreements"):
#             os.mkdir(f"outputs/{company.company_name}/Officer and Director Indemnification Agreements")

#         for indemnitee_name, indemnitee_address in indemnitees:
#             values = {
#                 "company_name": company.company_name,
#                 "formation_date": company.formation_date.strftime("%B %d, %Y"),
#                 "indemnitee_name": indemnitee_name,
#                 "company_address": company.company_address,
#                 "ceo_name": ceo_name,
#                 "indemnitee_address": indemnitee_address
#             }
#             if company.formation_state == "Delaware":
#                 indem_doc = docx.Document(de_indemnification_agreement_path)
#                 updated_doc = create_de_indemnification_agreement(doc=indem_doc, values=values)
#             if company.formation_state == "California":
#                 indem_doc = docx.Document(ca_indemnification_agreement_path)
#                 updated_doc = create_ca_indemnification_agreement(doc=indem_doc, values=values)
#             output_path = f"outputs/{company.company_name}/Officer and Director Indemnification Agreements/Indemnification_Agreement_{indemnitee_name}.docx"
#             updated_doc.save(output_path)

#     ### Create Individual and Corporate Stock Purchase Agreements
#     if "STOCK PURCHASE AGREEMENT" in download_checkbox:
#         if not os.path.exists(f"outputs/{company.company_name}/Stock Purchase Agreements"):
#             os.mkdir(f"outputs/{company.company_name}/Stock Purchase Agreements")

#         for stockholder_name, stockholder_type, stockholder_formation_state, stockholder_address, stockholder_spouse_name, num_shares, purchase_price, total_price, _, total_vesting, cliff, monthly_vesting in stockholders:
#             values = {
#                 "company_name": company.company_name,
#                 "formation_state": company.formation_state,
#                 "formation_date": company.formation_date.strftime("%B %d, %Y"),
#                 "purchaser_name": stockholder_name,
#                 "purchaser_share_purchased": num_shares,
#                 "purchase_price": purchase_price,
#                 "purchaser_total_price": total_price,
#                 "company_address": company.company_address,
#                 "purchaser_address": stockholder_address
#             }
#             if stockholder_type == EntityType.corporate:
#                 stock_purchase_agreement_doc = docx.Document(stock_purchase_agreement_entity_path)
#                 values["purchaser_formation_state"] = stockholder_formation_state
#                 values["stock_class"] = "Common"
#                 updated_doc = create_entity_stock_purchase_agreement(doc=stock_purchase_agreement_doc, values=values)
#             else:
#                 if (total_vesting and cliff and monthly_vesting):
#                     stock_purchase_agreement_doc = docx.Document(stock_purchase_agreement_person_vest_path)
#                     values["total_vesting_at_cliff"] = total_vesting
#                     values["cliff"] = cliff
#                     values["monthly_vesting_after_cliff"] = monthly_vesting
#                     values["ceo_name"] = ceo_name
#                     values["agreement_date"] = company.formation_date.strftime("%B %d, %Y")
#                     values["secretary_name"] = secretary_name
#                     values["spouse_name"] = stockholder_spouse_name
#                     updated_doc = create_natural_person_vesting_stock_purchase_agreement(doc=stock_purchase_agreement_doc, values=values)
#                 else:
#                     stock_purchase_agreement_doc = docx.Document(stock_purchase_agreement_person_no_vest_path)
#                     values["ceo_name"] = ceo_name
#                     updated_doc = create_natural_person_no_vesting_stock_purchase_agreement(doc=stock_purchase_agreement_doc, values=values)
#             output_path = f"outputs/{company.company_name}/Stock Purchase Agreements/Stock_Purchase_Agreement_{stockholder_name}.docx"
#             updated_doc.save(output_path)

#     ### Create Founder IP Assignment Agreements
#     if "IP CONTRIBUTION AND ASSIGNMENT AGREEMENT" in download_checkbox:
#         if not os.path.exists(f"outputs/{company.company_name}/IP Assignment Agreements"):
#             os.mkdir(f"outputs/{company.company_name}/IP Assignment Agreements")

#         for founder_name, founder_address, num_shares, purchase_price, total_price in founders:
#             ip_assignment_agreement_doc = docx.Document(ip_assignment_agreement_path)
#             values = {
#                 "company_name": company.company_name,
#                 "formation_state": company.formation_state,
#                 "formation_date": company.formation_date.strftime("%B %d, %Y"),
#                 "purchaser_name": founder_name,
#                 "purchaser_share_purchased": num_shares,
#                 "purchase_price": purchase_price,
#                 "purchaser_total_price": total_price,
#                 "ceo_name": ceo_name,
#                 "company_address": company.company_address,
#                 "purchaser_address": founder_address
#             }
#             updated_doc = create_ip_assignment_agreement(doc=ip_assignment_agreement_doc, values=values)
#             output_path = f"outputs/{company.company_name}/IP Assignment Agreements/IP_Agreement_{founder_name}.docx"
#             updated_doc.save(output_path)

#     ### Create EIN
#     if "EMPLOYER IDENTIFICATION NUMBER (EIN) AUTHORIZATION" in download_checkbox:
#         ein_doc = docx.Document(ein_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_date": company.formation_date.strftime("%B %d, %Y"),
#             "formation_state": company.formation_state,
#             "ceo_name": ceo_name
#         }
#         updated_doc = create_ein_application_authorization(doc=ein_doc, values=values)
#         output_path = f"outputs/{company.company_name}/Employer Identification Number Authorization.docx"
#         updated_doc.save(output_path)

#     ### Create Stock Certificates
#     if "STOCK CERTIFICATE" in download_checkbox:
#         stock_certificate_doc = docx.Document(stock_certificate_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "num_common_stock": company.num_common_stock,
#             "par_value": f"${company.par_value:.5f}",
#         }
#         updated_doc = create_stock_certificate(doc=stock_certificate_doc, values=values)
#         output_path = f"outputs/{company.company_name}/Stock_Certificate.docx"
#         updated_doc.save(output_path)

#     ### Create Stockholder Indemnification Agreement
#     if "STOCKHOLDER INDEMNIFICATION AGREEMENT" in download_checkbox:
#         stockholder_indemnification_agreement_doc = docx.Document(stockholder_indemnification_agreement_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "execution_date": (company.formation_date + timedelta(days=30)).strftime("%B %d, %Y")
#         }
#         updated_doc = create_stockholder_consent_indemnification_agreement(doc=stockholder_indemnification_agreement_doc, values=values, directors=directors)
#         output_path = f"outputs/{company.company_name}/Stockholder Indemnification Agreement.docx"
#         updated_doc.save(output_path)

#     ### Create Stock Incentive Plan
#     if "STOCK INCENTIVE PLAN" in download_checkbox:
#         stock_incentive_plan_doc = docx.Document(stock_incentive_plan_path)
#         values = {
#             "formation_state": company.formation_state,
#             "current_year": datetime.today().year,
#             "num_reserved_shares": company.num_reserved_stock,
#             "company_name": company.company_name
#         }
#         updated_doc = create_stock_incentive_plan(doc=stock_incentive_plan_doc, values=values)
#         output_path = f"outputs/{company.company_name}/Stock Incentive Plan.docx"
#         updated_doc.save(output_path)

#     ### Create Board Consent - Adopt Stock Plan
#     if "BOARD CONSENT ADOPT STOCK PLAN" in download_checkbox:
#         board_consent_adopt_stock_plan_doc = docx.Document(board_consent_adopt_stock_plan_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "current_year": datetime.today().year,
#             "num_reserved_shares": company.num_reserved_stock,
#             "formation_date": company.formation_date.strftime("%B %d, %Y")
#         }
#         updated_doc = create_board_consent_adopt_stock_plan(doc=board_consent_adopt_stock_plan_doc, values=values, directors=directors)
#         output_path = f"outputs/{company.company_name}/Board Consent Adopt Stock Plan.docx"
#         updated_doc.save(output_path)

#     ### Create Stockholder Consent - Adopt Stock Plan
#     if "STOCKHOLDER CONSENT ADOPT STOCK PLAN" in download_checkbox:
#         stockholder_consent_adopt_stock_plan_doc = docx.Document(stockholder_consent_adopt_stock_plan_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "current_year": datetime.today().year,
#             "num_reserved_shares": company.num_reserved_stock,
#             "execution_date": (company.formation_date + timedelta(days=30)).strftime("%B %d, %Y")
#         }
#         updated_doc = create_stockholder_consent_adopt_stock_plan(doc=stockholder_consent_adopt_stock_plan_doc, values=values, directors=directors)
#         output_path = f"outputs/{company.company_name}/Stockholder Consent Adopt Stock Plan.docx"
#         updated_doc.save(output_path)

#     ### Create Secretary Certificate Adopt Bylaws
#     if "SECRETARY CERTIFICATE ADOPT BYLAWS" in download_checkbox:
#         secretary_certificate_adopt_bylaws_doc = docx.Document(secretary_certificate_adopt_bylaws_path)
#         values = {
#             "company_name": company.company_name,
#             "formation_state": company.formation_state,
#             "formation_date": company.formation_date.strftime("%B %d, %Y"),
#             "secretary_name": secretary_name
#         }
#         updated_doc = create_secretary_certificate_adopt_bylaws(doc=secretary_certificate_adopt_bylaws_doc, values=values)
#         output_path = f"outputs/{company.company_name}/Secretary Certificate Adopt Bylaws.docx"
#         updated_doc.save(output_path)

#     zip_path = f"outputs/{company.company_name}.zip"
#     with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
#         directory_path = f"outputs/{company.company_name}"
#         for root, dirs, files in os.walk(directory_path):
#             for file in files:
#                 file_path = os.path.join(root, file)
#                 zipf.write(file_path, os.path.relpath(file_path, directory_path))

#     return gr.FileExplorer(root_dir=".")


# def update_root_dir():
#     return gr.FileExplorer(root_dir="./outputs/")


# def next_tab(tab_id: int):
#     return gr.Tabs(selected=tab_id + 1)


# def fill_ss4(*args):
#     (
#         ss4_1, ss4_2, ss4_3, ss4_4a, ss4_4b, ss4_5a, ss4_5b, ss4_6, ss4_7a, ss4_7b, ss4_8a, ss4_8b, ss4_8c,
#         ss4_9aSoleProprietor, ss4_9aSoleProprietorINFO, ss4_9aEstate, ss4_9aEstateINFO, ss4_9aPartnership,
#         ss4_9aPlanAdministrator, ss4_9aPlanAdministratorINFO, ss4_9aCorporation, ss4_9aCorporationINFO,
#         ss4_9aTrust, ss4_9aTrustINFO, ss4_9aPersonalServiceCorporation, ss4_9aMilitary, ss4_9aStateGovernment,
#         ss4_9aChurch, ss4_9aFarmers, ss4_9aFederalGovernment, ss4_9aOtherNonprofitOrganization, 
#         ss4_9aOtherNonprofitOrganizationINFO, ss4_9aREMIC, ss4_9aIndianTribalGovernments, ss4_9aOther,
#         ss4_9aOtherINFO, ss4_9aGEN, ss4_9bState, ss4_9bForeign, ss4_10Banking, ss4_10BankingINFO, 
#         ss4_10StartNewBusiness, ss4_10StartNewBusinessINFO, ss4_10ChangedOrganization, ss4_10ChangedOrganizationINFO,
#         ss4_10PurchasedBusiness, ss4_10HiredEmployees, ss4_10CreatedTrust, ss4_10CreatedTrustINFO, 
#         ss4_10ComplianceIRS, ss4_10PensionPlan, ss4_10PensionPlanINFO, ss4_10Others, ss4_10OthersINFO, 
#         ss4_11, ss4_12, ss4_13Agricultural, ss4_13Household, ss4_13Other, ss4_14, ss4_15, ss4_16HealthCare,
#         ss4_16WholesaleAgent, ss4_16Construction, ss4_16Rental, ss4_16Transporting, ss4_16Accommodation, 
#         ss4_16WholesaleOther, ss4_16Retail, ss4_16RealEstate, ss4_16Manufacturing, ss4_16Finance, 
#         ss4_16Other, ss4_16OtherINFO, ss4_17, ss4_18, ss4_18EIN, ss4_DesigneeName, ss4_DesigneeTelephone, 
#         ss4_DesigneeAddress, ss4_DesigneeFax, ss4_ApplicantName, ss4_ApplicantTelephone, ss4_ApplicantFax
#     ) = args
#     pdf = PdfReader(ss4_path)
#     values = {
#         '(EIN)': "", 
#         '(1)': ss4_1, 
#         '(2)': ss4_2, 
#         '(3)': ss4_3, 
#         '(4a)': ss4_4a, 
#         '(4b)': ss4_4b, 
#         '(5a)': ss4_5a, 
#         '(5b)': ss4_5b, 
#         '(6)': ss4_6, 
#         '(7a)': ss4_7a, 
#         '(7b)': ss4_7b, 
#         '(8aYes)': True if ss4_8a == "Yes" else False, 
#         '(8aNo)': True if ss4_8a == "No" else False, 
#         '(8b)': ss4_8b, 
#         '(8cYes)': True if ss4_8c == "Yes" else False, 
#         '(8cNo)': True if ss4_8c == "No" else False, 
#         '(9aSoleProprietor)': ss4_9aSoleProprietor, 
#         '(9aSoleProprietorINFO)': ss4_9aSoleProprietorINFO, 
#         '(9aEstate)': ss4_9aEstate, 
#         '(9aEstateINFO)': ss4_9aEstateINFO, 
#         '(9aPartnership)': ss4_9aPartnership, 
#         '(9aPlanAdministrator)': ss4_9aPlanAdministrator, 
#         '(9aPlanAdministratorINFO)': ss4_9aPlanAdministratorINFO, 
#         '(9aCorporation)': ss4_9aCorporation, 
#         '(9aCorporationINFO)': ss4_9aCorporationINFO, 
#         '(9aTrust)': ss4_9aTrust, 
#         '(9aTrustINFO)': ss4_9aTrustINFO, 
#         '(9aPersonalServiceCorporation)': ss4_9aPersonalServiceCorporation, 
#         '(9aMilitary)': ss4_9aMilitary, 
#         '(9aStateGovernment)': ss4_9aStateGovernment, 
#         '(9aChurch)': ss4_9aChurch, 
#         '(9aFarmers)': ss4_9aFarmers, 
#         '(9aFederalGovernment)': ss4_9aFederalGovernment, 
#         '(9aOtherNonprofitOrganization)': ss4_9aOtherNonprofitOrganization, 
#         '(9aOtherNonprofitOrganizationINFO)': ss4_9aOtherNonprofitOrganizationINFO, 
#         '(9aREMIC)': ss4_9aREMIC, 
#         '(9aIndianTribalGovernments)': ss4_9aIndianTribalGovernments, 
#         '(9aOther)': ss4_9aOther, 
#         '(9aOtherINFO)': ss4_9aOtherINFO, 
#         '(9aGEN)': ss4_9aGEN, 
#         '(9bState)': ss4_9bState, 
#         '(9bForeign)': ss4_9bForeign, 
#         '(10Banking)': ss4_10Banking, 
#         '(10BankingINFO)': ss4_10BankingINFO, 
#         '(10StartNewBusiness)': ss4_10StartNewBusiness, 
#         # '(10StartNewBusinessINFO)': ss4_10StartNewBusinessINFO, 
#         '(10StartNewBusinessINFO2)': ss4_10StartNewBusinessINFO, 
#         '(10ChangedOrganization)': ss4_10ChangedOrganization, 
#         '(10ChangedOrganizationINFO)': ss4_10ChangedOrganizationINFO, 
#         '(10PurchasedBusiness)': ss4_10PurchasedBusiness, 
#         '(10HiredEmployees)': ss4_10HiredEmployees, 
#         '(10CreatedTrust)': ss4_10CreatedTrust, 
#         '(10CreatedTrustINFO)': ss4_10CreatedTrustINFO, 
#         '(10ComplianceIRS)': ss4_10ComplianceIRS, 
#         '(10PensionPlan)': ss4_10PensionPlan, 
#         '(10PensionPlanINFO)': ss4_10PensionPlanINFO, 
#         '(10Others)': ss4_10Others, 
#         '(10OthersINFO)': ss4_10OthersINFO, 
#         '(11)': ss4_11, 
#         '(12)': ss4_12, 
#         '(13Agricultural)': ss4_13Agricultural, 
#         '(13Household)': ss4_13Household, 
#         '(13Other)': ss4_13Other, 
#         '(14)': ss4_14, 
#         '(15)': ss4_15, 
#         '(16HealthCare)': ss4_16HealthCare, 
#         '(16WholesaleAgent)': ss4_16WholesaleAgent, 
#         '(16Construction)': ss4_16Construction, 
#         '(16Rental)': ss4_16Rental, 
#         '(16Transporting)': ss4_16Transporting, 
#         '(16Accommodation)': ss4_16Accommodation, 
#         '(16WholesaleOther)': ss4_16WholesaleOther, 
#         '(16Retail)': ss4_16Retail, 
#         '(16RealEstate)': ss4_16RealEstate, 
#         '(16Manufacturing)': ss4_16Manufacturing, 
#         '(16Finance)': ss4_16Finance, 
#         '(16Other)': ss4_16Other, 
#         '(16OtherINFO)': ss4_16OtherINFO, 
#         '(17)': ss4_17, 
#         '(18Yes)': True if ss4_18 == "Yes" else False, 
#         '(18No)': True if ss4_18 == "No" else False, 
#         '(18EIN)': ss4_18EIN, 
#         '(DesigneeName)': ss4_DesigneeName, 
#         '(DesigneeTelephone)': ss4_DesigneeTelephone, 
#         '(DesigneeAddress)': ss4_DesigneeAddress, 
#         '(DesigneeFax)': ss4_DesigneeFax, 
#         '(ApplicantName)': ss4_ApplicantName, 
#         '(ApplicantTelephone)': ss4_ApplicantTelephone, 
#         '(ApplicantFax)': ss4_ApplicantFax
#     }
#     updated_pdf = create_ss4(pdf=pdf, values=values)
#     output_path = f"outputs/SS4.pdf"
#     PdfWriter().write(output_path, updated_pdf)
#     return output_path


# def create_invoice(base_price: int, agent_states: List[str], regi_states: List[str], us_address_req: str, us_address_value: int, sop_reserve_req: str, sop_reserve_value: int, add_founder: int, add_founder_value: int, docsign_req: str, docsign_value: int, *args):
#     ### ARGS
#     ### 0 - 49 -> AGENT STATE VALUES
#     ### 50 - 99 -> REGISTRATION STATE VALUES

#     agent_values = [i for i in args[:50] if i]
#     regi_values = [i for i in args[50:] if i]
#     sop_reserve_req = 1 if sop_reserve_req == "Yes" else 0
#     us_address_req = 1 if us_address_req == "Yes" else 0
#     docsign_req = 1 if docsign_req == "Yes" else 0
#     add_founder = add_founder - 1

#     try:
#         workbook = load_workbook(invoice_path)
#     except FileNotFoundError:
#         workbook = Workbook()

#     sheet_name = 'Company_Itemised_invoice'
#     if sheet_name in workbook.sheetnames:
#         sheet = workbook[sheet_name]
#     else:
#         sheet = workbook.create_sheet(sheet_name)

#     agent_count = 0
#     for i, state in enumerate(agent_states):
#         agent_count+=1
#         sheet.cell(row=i+5, column=6, value=state)

#     for i, value in enumerate(agent_values):
#         sheet.cell(row=i+5, column=8, value=value)

#     regi_count = 0
#     for i, state in enumerate(regi_states):
#         regi_count+=1
#         sheet.cell(row=i+55, column=6, value=state)

#     for i, value in enumerate(regi_values):
#         sheet.cell(row=i+55, column=8, value=value)

#     cell_mapping = {
#         "H3": base_price,
#         "G105": us_address_req,
#         "H105": us_address_value,
#         "G106": sop_reserve_req,
#         "H106": sop_reserve_value,
#         "G107": add_founder,
#         "H107": add_founder_value,
#         "G108": docsign_req,
#         "H108": docsign_value
#     }
#     for cell, value in cell_mapping.items():
#         sheet[cell] = value

#     sheet.delete_rows(5 + agent_count, 50 - agent_count)
#     sheet.delete_rows(5 + agent_count + regi_count, 50 - regi_count)

#     for row in range(5, 5 + agent_count + regi_count + 4):
#         sheet[f'I{row}'] = f'=G{row}*H{row}'

#     sum_formula = f'=SUM(I3:I{3+ agent_count + regi_count + 4})'
#     sheet[f'I{3 + 1 + agent_count + regi_count + 5}'] = sum_formula

#     sheet.merge_cells(range_string=f"E5:E{5 + agent_count - 1}")
#     sheet.merge_cells(range_string=f"D5:D{5 + agent_count - 1}")
#     sheet.merge_cells(range_string=f"E{5 + agent_count}:E{5 + agent_count + regi_count - 1}")
#     sheet.merge_cells(range_string=f"D{5 + agent_count}:D{5 + agent_count + regi_count - 1}")
#     sheet.merge_cells(range_string=f"C5:C{5 + agent_count + regi_count + 2}")

#     output_path = f"outputs/Invoice.xlsx"
#     workbook.save(output_path)
#     return output_path


def fill_ss4_1(*args):
    # Unpack the arguments
    (
        ss4_1, ss4_2, ss4_3, ss4_4a, ss4_4b, ss4_5a, ss4_5b, ss4_6, ss4_7a, ss4_7b, ss4_8a, ss4_8b, ss4_8c,
        ss4_9aSoleProprietor, ss4_9aSoleProprietorINFO, ss4_9aEstate, ss4_9aEstateINFO, ss4_9aPartnership,
        ss4_9aPlanAdministrator, ss4_9aPlanAdministratorINFO, ss4_9aCorporation, ss4_9aCorporationINFO,
        ss4_9aTrust, ss4_9aTrustINFO, ss4_9aPersonalServiceCorporation, ss4_9aMilitary, ss4_9aStateGovernment,
        ss4_9aChurch, ss4_9aFarmers, ss4_9aFederalGovernment, ss4_9aOtherNonprofitOrganization, 
        ss4_9aOtherNonprofitOrganizationINFO, ss4_9aREMIC, ss4_9aIndianTribalGovernments, ss4_9aOther,
        ss4_9aOtherINFO, ss4_9aGEN, ss4_9bState, ss4_9bForeign, ss4_10Banking, ss4_10BankingINFO, 
        ss4_10StartNewBusiness, ss4_10StartNewBusinessINFO, ss4_10ChangedOrganization, ss4_10ChangedOrganizationINFO,
        ss4_10PurchasedBusiness, ss4_10HiredEmployees, ss4_10CreatedTrust, ss4_10CreatedTrustINFO, 
        ss4_10ComplianceIRS, ss4_10PensionPlan, ss4_10PensionPlanINFO, ss4_10Others, ss4_10OthersINFO, 
        ss4_11, ss4_12, ss4_13Agricultural, ss4_13Household, ss4_13Other, ss4_14, ss4_15, ss4_16HealthCare,
        ss4_16WholesaleAgent, ss4_16Construction, ss4_16Rental, ss4_16Transporting, ss4_16Accommodation, 
        ss4_16WholesaleOther, ss4_16Retail, ss4_16RealEstate, ss4_16Manufacturing, ss4_16Finance, 
        ss4_16Other, ss4_16OtherINFO, ss4_17, ss4_18, ss4_18EIN, ss4_DesigneeName, ss4_DesigneeTelephone, 
        ss4_DesigneeAddress, ss4_DesigneeFax, ss4_ApplicantName, ss4_ApplicantTelephone, ss4_ApplicantFax
    ) = args

    # Establish MySQL connection
    connection = mysql.connector.connect(
    host="sql3.freesqldatabase.com",
    port="3306",
    user="sql3716747",
    password="KuUTTfiiKQ",
    database="sql3716747"
    )

    # Create a cursor object using the connection
    cursor = connection.cursor()

    # Insert query with placeholders
    insert_query = """
    INSERT INTO ss4_table (
        ss4_1, ss4_2, ss4_3, ss4_4a, ss4_4b, ss4_5a, ss4_5b, ss4_6, ss4_7a, ss4_7b, ss4_8a, ss4_8b, ss4_8c, 
        ss4_9aSoleProprietor, ss4_9aSoleProprietorINFO, ss4_9aEstate, ss4_9aEstateINFO, ss4_9aPartnership, ss4_9aPlanAdministrator, 
        ss4_9aPlanAdministratorINFO, ss4_9aCorporation, ss4_9aCorporationINFO, ss4_9aTrust, ss4_9aTrustINFO,ss4_9aPersonalServiceCorporation, 
        ss4_9aMilitary, ss4_9aStateGovernment, ss4_9aChurch, ss4_9aFarmers, ss4_9aFederalGovernment, ss4_9aOtherNonprofitOrganization, 
        ss4_9aOtherNonprofitOrganizationINFO, ss4_9aREMIC, ss4_9aIndianTribalGovernments, ss4_9aOther, ss4_9aOtherINFO, ss4_9aGEN, ss4_9bState, 
        ss4_9bForeign, ss4_10Banking, ss4_10BankingINFO, ss4_10StartNewBusiness, ss4_10StartNewBusinessINFO, ss4_10ChangedOrganization, 
        ss4_10ChangedOrganizationINFO, ss4_10PurchasedBusiness, ss4_10HiredEmployees, ss4_10CreatedTrust, ss4_10CreatedTrustINFO, ss4_10ComplianceIRS, 
        ss4_10PensionPlan, ss4_10PensionPlanINFO, ss4_10Others, ss4_10OthersINFO, ss4_11, ss4_12, ss4_13Agricultural, ss4_13Household, 
        ss4_13Other, ss4_14, ss4_15, ss4_16HealthCare, ss4_16WholesaleAgent, ss4_16Construction, ss4_16Rental, ss4_16Transporting, 
        ss4_16Accommodation, ss4_16WholesaleOther, ss4_16Retail, ss4_16RealEstate, ss4_16Manufacturing, ss4_16Finance, ss4_16Other, ss4_16OtherINFO, ss4_17, 
        ss4_18, ss4_18EIN, ss4_DesigneeName, ss4_DesigneeTelephone, ss4_DesigneeAddress, ss4_DesigneeFax, ss4_ApplicantName, ss4_ApplicantTelephone, ss4_ApplicantFax
    ) 
    VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
        %s, %s, %s, %s
    )
    """

    data = (
        ss4_1, ss4_2, ss4_3, ss4_4a, ss4_4b, ss4_5a, ss4_5b, ss4_6, ss4_7a, ss4_7b, ss4_8a, ss4_8b, ss4_8c,
        ss4_9aSoleProprietor, ss4_9aSoleProprietorINFO, ss4_9aEstate, ss4_9aEstateINFO, ss4_9aPartnership,
        ss4_9aPlanAdministrator, ss4_9aPlanAdministratorINFO, ss4_9aCorporation, ss4_9aCorporationINFO,
        ss4_9aTrust, ss4_9aTrustINFO, ss4_9aPersonalServiceCorporation, ss4_9aMilitary, ss4_9aStateGovernment,
        ss4_9aChurch, ss4_9aFarmers, ss4_9aFederalGovernment, ss4_9aOtherNonprofitOrganization, 
        ss4_9aOtherNonprofitOrganizationINFO, ss4_9aREMIC, ss4_9aIndianTribalGovernments, ss4_9aOther,
        ss4_9aOtherINFO, ss4_9aGEN, ss4_9bState, ss4_9bForeign, ss4_10Banking, ss4_10BankingINFO, 
        ss4_10StartNewBusiness, ss4_10StartNewBusinessINFO, ss4_10ChangedOrganization, ss4_10ChangedOrganizationINFO,
        ss4_10PurchasedBusiness, ss4_10HiredEmployees, ss4_10CreatedTrust, ss4_10CreatedTrustINFO, 
        ss4_10ComplianceIRS, ss4_10PensionPlan, ss4_10PensionPlanINFO, ss4_10Others, ss4_10OthersINFO, 
        ss4_11, ss4_12, ss4_13Agricultural, ss4_13Household, ss4_13Other, ss4_14, ss4_15, ss4_16HealthCare,
        ss4_16WholesaleAgent, ss4_16Construction, ss4_16Rental, ss4_16Transporting, ss4_16Accommodation, 
        ss4_16WholesaleOther, ss4_16Retail, ss4_16RealEstate, ss4_16Manufacturing, ss4_16Finance, 
        ss4_16Other, ss4_16OtherINFO, ss4_17, ss4_18, ss4_18EIN, ss4_DesigneeName, ss4_DesigneeTelephone, 
        ss4_DesigneeAddress, ss4_DesigneeFax, ss4_ApplicantName, ss4_ApplicantTelephone, ss4_ApplicantFax
    )

    try:
        # Execute the query
        cursor.execute(insert_query, data)
        
        # Commit the transaction
        connection.commit()
        print("Data inserted successfully into ss4_table.")
    
    except mysql.connector.Error as error:
        print(f"Failed to insert data into MySQL table: {error}")

    finally:
        # Close cursor and connection
        cursor.close()
        connection.close()





    pdf = PdfReader(ss4_path)
    values = {
        '(EIN)': "", 
        '(1)': ss4_1, 
        '(2)': ss4_2, 
        '(3)': ss4_3, 
        '(4a)': ss4_4a, 
        '(4b)': ss4_4b, 
        '(5a)': ss4_5a, 
        '(5b)': ss4_5b, 
        '(6)': ss4_6, 
        '(7a)': ss4_7a, 
        '(7b)': ss4_7b, 
        '(8aYes)': True if ss4_8a == "Yes" else False, 
        '(8aNo)': True if ss4_8a == "No" else False, 
        '(8b)': ss4_8b, 
        '(8cYes)': True if ss4_8c == "Yes" else False, 
        '(8cNo)': True if ss4_8c == "No" else False, 
        '(9aSoleProprietor)': ss4_9aSoleProprietor, 
        '(9aSoleProprietorINFO)': ss4_9aSoleProprietorINFO, 
        '(9aEstate)': ss4_9aEstate, 
        '(9aEstateINFO)': ss4_9aEstateINFO, 
        '(9aPartnership)': ss4_9aPartnership, 
        '(9aPlanAdministrator)': ss4_9aPlanAdministrator, 
        '(9aPlanAdministratorINFO)': ss4_9aPlanAdministratorINFO, 
        '(9aCorporation)': ss4_9aCorporation, 
        '(9aCorporationINFO)': ss4_9aCorporationINFO, 
        '(9aTrust)': ss4_9aTrust, 
        '(9aTrustINFO)': ss4_9aTrustINFO, 
        '(9aPersonalServiceCorporation)': ss4_9aPersonalServiceCorporation, 
        '(9aMilitary)': ss4_9aMilitary, 
        '(9aStateGovernment)': ss4_9aStateGovernment, 
        '(9aChurch)': ss4_9aChurch, 
        '(9aFarmers)': ss4_9aFarmers, 
        '(9aFederalGovernment)': ss4_9aFederalGovernment, 
        '(9aOtherNonprofitOrganization)': ss4_9aOtherNonprofitOrganization, 
        '(9aOtherNonprofitOrganizationINFO)': ss4_9aOtherNonprofitOrganizationINFO, 
        '(9aREMIC)': ss4_9aREMIC, 
        '(9aIndianTribalGovernments)': ss4_9aIndianTribalGovernments, 
        '(9aOther)': ss4_9aOther, 
        '(9aOtherINFO)': ss4_9aOtherINFO, 
        '(9aGEN)': ss4_9aGEN, 
        '(9bState)': ss4_9bState, 
        '(9bForeign)': ss4_9bForeign, 
        '(10Banking)': ss4_10Banking, 
        '(10BankingINFO)': ss4_10BankingINFO, 
        '(10StartNewBusiness)': ss4_10StartNewBusiness, 
        # '(10StartNewBusinessINFO)': ss4_10StartNewBusinessINFO, 
        '(10StartNewBusinessINFO2)': ss4_10StartNewBusinessINFO, 
        '(10ChangedOrganization)': ss4_10ChangedOrganization, 
        '(10ChangedOrganizationINFO)': ss4_10ChangedOrganizationINFO, 
        '(10PurchasedBusiness)': ss4_10PurchasedBusiness, 
        '(10HiredEmployees)': ss4_10HiredEmployees, 
        '(10CreatedTrust)': ss4_10CreatedTrust, 
        '(10CreatedTrustINFO)': ss4_10CreatedTrustINFO, 
        '(10ComplianceIRS)': ss4_10ComplianceIRS, 
        '(10PensionPlan)': ss4_10PensionPlan, 
        '(10PensionPlanINFO)': ss4_10PensionPlanINFO, 
        '(10Others)': ss4_10Others, 
        '(10OthersINFO)': ss4_10OthersINFO, 
        '(11)': ss4_11, 
        '(12)': ss4_12, 
        '(13Agricultural)': ss4_13Agricultural, 
        '(13Household)': ss4_13Household, 
        '(13Other)': ss4_13Other, 
        '(14)': ss4_14, 
        '(15)': ss4_15, 
        '(16HealthCare)': ss4_16HealthCare, 
        '(16WholesaleAgent)': ss4_16WholesaleAgent, 
        '(16Construction)': ss4_16Construction, 
        '(16Rental)': ss4_16Rental, 
        '(16Transporting)': ss4_16Transporting, 
        '(16Accommodation)': ss4_16Accommodation, 
        '(16WholesaleOther)': ss4_16WholesaleOther, 
        '(16Retail)': ss4_16Retail, 
        '(16RealEstate)': ss4_16RealEstate, 
        '(16Manufacturing)': ss4_16Manufacturing, 
        '(16Finance)': ss4_16Finance, 
        '(16Other)': ss4_16Other, 
        '(16OtherINFO)': ss4_16OtherINFO, 
        '(17)': ss4_17, 
        '(18Yes)': True if ss4_18 == "Yes" else False, 
        '(18No)': True if ss4_18 == "No" else False, 
        '(18EIN)': ss4_18EIN, 
        '(DesigneeName)': ss4_DesigneeName, 
        '(DesigneeTelephone)': ss4_DesigneeTelephone, 
        '(DesigneeAddress)': ss4_DesigneeAddress, 
        '(DesigneeFax)': ss4_DesigneeFax, 
        '(ApplicantName)': ss4_ApplicantName, 
        '(ApplicantTelephone)': ss4_ApplicantTelephone, 
        '(ApplicantFax)': ss4_ApplicantFax
    }
    updated_pdf = create_ss4(pdf=pdf, values=values)
    output_path = f"outputs/SS4.pdf"
    PdfWriter().write(output_path, updated_pdf)
    return output_path











with gr.Blocks(title="Questionnaire") as demo:
    company_elems = []
    entity_elems = []
    role_elems = []
    stockholder_elems = []
    retrieved_elems = []
    ss4_elems = []
    
    entity_ids = []
    role_ids = []
    stockholder_ids = []

    with gr.Tabs() as tabs:

        # ### COMPANY DETAILS

        # with gr.Tab(label="1. Add Company Details", id=0):
        #     with gr.Row(visible=VISIBLE):
        #         tab0_id = gr.Number(label="Tab ID", interactive=False, value=0)

        #     with gr.Row():
        #         gr.Markdown("## Information about the company.")

        #     with gr.Row(variant="panel"):
        #         new_or_existing = gr.Radio(show_label=False, info="Create New or Choose Existing Form", choices=["New", "Existing"], value="New")
        #         existing_dropdown = gr.Dropdown(show_label=False, choices=[], allow_custom_value=True, interactive=True, visible=False)

        #         def enable_existing_dropdown(new_or_existing: str):
        #             if new_or_existing == "Existing":
        #                 with Session() as session:
        #                     company_info_list = session.query(Company.company_id, Company.company_name).all()
        #                     company_names = [f"{company.company_id} - {company.company_name}" for company in company_info_list]
        #                 return gr.Dropdown(visible=True, choices=company_names)
        #             return gr.Dropdown(visible=False)
        #         new_or_existing.input(fn=enable_existing_dropdown, inputs=new_or_existing, outputs=existing_dropdown)
                    
        #     with gr.Row(visible=VISIBLE):
        #         company_id = gr.Number(label="Company ID", interactive=False)

        #         def reset(new_or_existing: str):
        #             if new_or_existing == "New":
        #                 return gr.Number(value=0)
        #         new_or_existing.input(fn=reset, inputs=new_or_existing, outputs=company_id)

        #     with gr.Row(variant="panel"):
        #         formation_state = gr.Radio(label="State of Formation", choices=["Delaware", "California"], value="Delaware", info="Which state do you wish to establish your new company?")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             company_name = gr.Textbox(label="Company Name", info="Enter the desired name of your new company.")
        #         with gr.Column():
        #             company_address = gr.TextArea(label="Company Address", info="What is the new company’s United States address (temporary is ok)?")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             agent_name = gr.Textbox(label="Agent Name", info="Enter the name of your registered agent.")
        #         with gr.Column():
        #             agent_address = gr.TextArea(label="Agent Address", info="Enter the address of your registered agent.")
        #         with gr.Column():
        #             agent_city = gr.Textbox(label="Agent City")
        #             agent_county = gr.Textbox(label="Agent County")
        #             agent_zipcode = gr.Textbox(label="Agent Zipcode")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             par_value = gr.Number(label="Par Value", value=0.00001, step=0.00001, minimum=0)
        #         with gr.Column():
        #             num_common_stock = gr.Number(label="Number of Common Stocks", value=10000000, precision=0, minimum=0)
        #         with gr.Column():
        #             num_reserved_stock = gr.Number(label= "Number of Reserved Stocks", value=5000000, precision=0, minimum=0)

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             incorporator_name = gr.Textbox(label="Incorporator Name", info="Enter the name of the incorporator.")
        #         with gr.Column():
        #             incorporator_address = gr.TextArea(label="Incorporator Address", info="Enter the address of the incorporator.")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             fy_end = gr.Textbox(label="Fiscal Year End", value="December 31", info="What is the fiscal year of the new company? The fiscal year generally ends on December 31.")
        #         with gr.Column():
        #             submission_date = Calendar(label="Submission Date", value=datetime.today(), type="datetime")
        #         with gr.Column():
        #             formation_date = Calendar(label="Formation Date", value=datetime.today(), type="datetime")

        #     with gr.Row(variant="panel"):
        #         company_bank = gr.Textbox(label="Company Bank", info="The name of the bank the bank account will be opened in.")

        #     with gr.Row():
        #         gr.Markdown("Please review the entries before hitting the save button. The current version does not support any changes once these details are saved. Also, please ensure all entries are filled out to have the correct forms generated.")

        #     with gr.Row(variant="panel"):
        #         save_company_btn = gr.Button(value="Save", variant="primary")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             save_company_message = gr.Textbox(label="Message", interactive=False)
        #         with gr.Column():
        #             next_btn = gr.Button(value="Next")
        #             next_btn.click(fn=next_tab, inputs=tab0_id, outputs=tabs)

        #     company_elems.append(company_id)
        #     company_elems.append(formation_state)
        #     company_elems.append(company_name)
        #     company_elems.append(company_address)
        #     company_elems.append(agent_name)
        #     company_elems.append(agent_address)
        #     company_elems.append(agent_city)
        #     company_elems.append(agent_county)
        #     company_elems.append(agent_zipcode)
        #     company_elems.append(par_value)
        #     company_elems.append(num_common_stock)
        #     company_elems.append(num_reserved_stock)
        #     company_elems.append(incorporator_name)
        #     company_elems.append(incorporator_address)
        #     company_elems.append(fy_end)
        #     company_elems.append(submission_date)
        #     company_elems.append(formation_date)
        #     company_elems.append(company_bank)

        #     retrieved_elems.extend(company_elems[1:])

        #     save_company_btn.click(fn=save_company_data, inputs=company_elems, outputs=[company_id, save_company_message])

        # ### ENTITY DETAILS

        # with gr.Tab(label="2. Add Officers', Directors' & Stockholders' Details", id=1):
        #     with gr.Row(visible=VISIBLE):
        #         tab1_id = gr.Number(label="Tab ID", interactive=False, value=1)

        #     with gr.Row():
        #         gr.Markdown("## Information about company management, board and stockholders.")

        #     with gr.Row():
        #         gr.Markdown("Mention details of all unique members in the company, including any corporate stockholders")

        #     with gr.Row(variant="panel", visible=VISIBLE):
        #         num_entities = gr.Slider(label="Number of unique members in the company, including any corporate stockholders", info="What is the total number of unique founders, directors, officers, any individual and corporate stockholders will the new company have?", minimum=1, maximum=MAX_NUM_ENTITIES, step=1, value=MAX_NUM_ENTITIES)
        #         retrieved_elems.append(num_entities)

        #     entity_rows = []
        #     with gr.Row() as entity_row:
        #         with gr.Row():
        #             with gr.Row(visible=VISIBLE):
        #                 entity_id = gr.Number(label="Entity ID", interactive=False)

        #             with gr.Row(variant="panel"):
        #                 with gr.Column(scale=1):
        #                     entity_type = gr.Dropdown(label="Type", choices=["Individual", "Corporate"], value="Individual")
        #                     entity_formation_state = gr.Textbox(label="Formation State", visible=False)

        #                     def enable_entity_formation_state(entity_type: str):
        #                         return gr.Textbox(visible=True) if entity_type == "Corporate" else gr.Textbox(visible=False)
        #                     entity_type.change(fn=enable_entity_formation_state, inputs=entity_type, outputs=entity_formation_state)

        #                     entity_name = gr.Textbox(label="Name")
        #                     entity_addr = gr.TextArea(label="Address")
        #                 with gr.Column():
        #                     entity_spouse_name = gr.Textbox(label="Spouse Name")
        #                     entity_phone = gr.Textbox(label="Phone")
        #                     entity_email = gr.Textbox(label="Email")
        #                     entity_ssn = gr.Textbox(label="Social Security Number")

        #         entity_ids.append(entity_id)

        #         entity_elems.append(entity_id)
        #         entity_elems.append(entity_name)
        #         entity_elems.append(entity_addr)
        #         entity_elems.append(entity_phone)
        #         entity_elems.append(entity_email)
        #         entity_elems.append(entity_ssn)
        #         entity_elems.append(entity_spouse_name)
        #         entity_elems.append(entity_type)
        #         entity_elems.append(entity_formation_state)

        #         retrieved_elems.extend(entity_elems)

        #     entity_rows.append(entity_row)

        #     for i in range(MAX_NUM_ENTITIES-1):
        #         with gr.Row() as entity_row:
        #             with gr.Row():
        #                 with gr.Row(visible=VISIBLE):
        #                     _entity_id = gr.Number(label="Entity ID", interactive=False)
        #                 with gr.Row(variant="panel"):
        #                     with gr.Column():
        #                         _entity_type = gr.Dropdown(label="Type", choices=["Individual", "Corporate"], value="Individual")
        #                         _entity_formation_state = gr.Textbox(label="Formation State", visible=False)
        #                         _entity_type.change(fn=enable_entity_formation_state, inputs=_entity_type, outputs=_entity_formation_state)
        #                         _entity_name = gr.Textbox(label="Name")
        #                         _entity_addr = gr.TextArea(label="Address")
        #                     with gr.Column():
        #                         _entity_spouse_name = gr.Textbox(label="Spouse Name")
        #                         _entity_phone = gr.Textbox(label="Phone")
        #                         _entity_email = gr.Textbox(label="Email")
        #                         _entity_ssn = gr.Textbox(label="Social Security Number")

        #             entity_ids.append(_entity_id)

        #             entity_elems.append(_entity_id)
        #             entity_elems.append(_entity_name)
        #             entity_elems.append(_entity_addr)
        #             entity_elems.append(_entity_phone)
        #             entity_elems.append(_entity_email)
        #             entity_elems.append(_entity_ssn)
        #             entity_elems.append(_entity_spouse_name)
        #             entity_elems.append(_entity_type)
        #             entity_elems.append(_entity_formation_state)

        #             retrieved_elems.extend(entity_elems[-9:])

        #         entity_rows.append(entity_row)

        #     def enable_entity_rows(num_entities: int):
        #         entity_rows = []
        #         for i in range(num_entities):
        #             entity_rows.append(gr.Row(visible=True))
        #         for i in range(MAX_NUM_ENTITIES - num_entities):
        #             entity_rows.append(gr.Row(visible=False))
        #         return entity_rows
        #     num_entities.change(fn=enable_entity_rows, inputs=num_entities, outputs=entity_rows)

        #     with gr.Row():
        #         gr.Markdown("Please review the entries before hitting the save button. The current version does not support any changes once these details are saved. Also, please ensure all entries are filled out to have the correct forms generated.")

        #     with gr.Row():
        #         save_entity_btn = gr.Button(value="Save", variant="primary")

        #     with gr.Row():
        #         with gr.Column():
        #             save_entity_message = gr.Textbox(label="Message", interactive=False)
        #             save_entity_btn.click(fn=save_entity_data, inputs=[company_id] + entity_elems, outputs=entity_ids + [save_entity_message])
        #         with gr.Column():
        #             next_btn = gr.Button(value="Next")
        #             next_btn.click(fn=next_tab, inputs=tab1_id, outputs=tabs)

        # ### ROLE DETAILS

        # with gr.Tab(label="3. Tag Executives & Board Members", id=2):
        #     with gr.Row(visible=VISIBLE):
        #         tab2_id = gr.Number(label="Tab ID", interactive=False, value=2)

        #     with gr.Row():
        #         gr.Markdown("## Names of executives and board members")

        #     def update_dropdown(entity_id: int):
        #         entity_names = []
        #         with Session() as session:
        #             company_id = session.scalars(select(Entity.company_id).where(Entity.entity_id == entity_id)).first()
        #             if company_id:
        #                 entity_names = session.scalars(select(Entity.name).where(Entity.company_id == company_id)).all()
        #         return gr.Dropdown(choices=entity_names)

        #     with gr.Row(visible=VISIBLE):
        #         with gr.Column():
        #             founder_ids = gr.Textbox(label="Founder IDs", interactive=False)
        #         with gr.Column():
        #             director_ids = gr.Textbox(label="Director IDs", interactive=False)

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             founder_names = gr.Dropdown(label="Founders", info="Select all founders from the dropdown.", multiselect=True, choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=founder_names)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=founder_names)
        #         with gr.Column():
        #             director_names = gr.Dropdown(label="Directors", info="Select all directors from the dropdown.", multiselect=True, choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=director_names)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=director_names)

        #     with gr.Row(variant="panel"):
        #         with gr.Column(visible=VISIBLE):
        #             chairman_id = gr.Number(label="Officer ID", interactive=False)
        #         with gr.Column():
        #             chairman_name = gr.Dropdown(label="Chairman", choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=chairman_name)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=chairman_name)

        #     with gr.Row(variant="panel"):
        #         with gr.Column(visible=VISIBLE):
        #             ceo_id = gr.Number(label="Officer ID", interactive=False)
        #         with gr.Column(scale=3):
        #             ceo_name = gr.Dropdown(label="CEO", choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=ceo_name)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=ceo_name)
        #         with gr.Column(scale=1):
        #             ceo_open_account = gr.Checkbox(label="Can Open Bank Account")
        #             ceo_draw_account = gr.Checkbox(label="Can Draw from Bank Account")

        #     with gr.Row(variant="panel"):
        #         with gr.Column(visible=VISIBLE):
        #             cfo_id = gr.Number(label="Officer ID", interactive=False)
        #         with gr.Column(scale=3):
        #             cfo_name = gr.Dropdown(label="CFO", choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=cfo_name)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=cfo_name)
        #         with gr.Column(scale=1):
        #             cfo_open_account = gr.Checkbox(label="Can Open Bank Account")
        #             cfo_draw_account = gr.Checkbox(label="Can Draw from Bank Account")

        #     with gr.Row(variant="panel"):
        #         with gr.Column(visible=VISIBLE):
        #             secretary_id = gr.Number(label="Officer ID", interactive=False)
        #         with gr.Column(scale=3):
        #             secretary_name = gr.Dropdown(label="Secretary", choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=secretary_name)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=secretary_name)
        #         with gr.Column(scale=1):
        #             secretary_open_account = gr.Checkbox(label="Can Open Bank Account")
        #             secretary_draw_account = gr.Checkbox(label="Can Draw from Bank Account")

        #     with gr.Row():
        #         gr.Markdown("Enter the positions and names of other executives you wish to appoint.")

        #     with gr.Row(visible=VISIBLE):
        #         num_other_officers = gr.Slider(label="Number of Other Executives", info="Enter the positions and names of other executives you wish to appoint.", minimum=0, maximum=MAX_NUM_OTHER_OFFICERS, value=MAX_NUM_OTHER_OFFICERS, step=1)

        #     role_ids.append(founder_ids)
        #     role_ids.append(director_ids)
        #     role_ids.append(chairman_id)
        #     role_ids.append(ceo_id)
        #     role_ids.append(cfo_id)
        #     role_ids.append(secretary_id)

        #     retrieved_elems.extend([
        #         founder_ids,
        #         director_ids,
        #         founder_names,
        #         director_names,
        #         chairman_id,
        #         chairman_name,
        #         ceo_id,
        #         ceo_name,
        #         ceo_open_account,
        #         ceo_draw_account,
        #         cfo_id,
        #         cfo_name,
        #         cfo_open_account,
        #         cfo_draw_account,
        #         secretary_id,
        #         secretary_name,
        #         secretary_open_account,
        #         secretary_draw_account,
        #         num_other_officers
        #     ])

        #     other_officer_rows = []
        #     for i in range(MAX_NUM_OTHER_OFFICERS):
        #         with gr.Row(variant="panel") as other_officer_row:
        #             with gr.Column(visible=VISIBLE):
        #                 officer_id = gr.Number(label="Officer ID", interactive=False)
        #             with gr.Column(scale=2):
        #                 officer_title = gr.Textbox(label="Title")
        #             with gr.Column(scale=2):
        #                 officer_name = gr.Dropdown(label="Name", choices=[], allow_custom_value=True)
        #                 save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=officer_name)
        #                 entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=officer_name)
        #             with gr.Column(scale=1):
        #                 officer_open_account = gr.Checkbox(label="Can Open Bank Account")
        #                 officer_draw_account = gr.Checkbox(label="Can Draw from Bank Account")
                
        #         role_ids.append(officer_id)

        #         role_elems.append(officer_id)
        #         role_elems.append(officer_title)
        #         role_elems.append(officer_name)
        #         role_elems.append(officer_open_account)
        #         role_elems.append(officer_draw_account)

        #         other_officer_rows.append(other_officer_row)
        #     retrieved_elems.extend(role_elems)

        #     def enable_other_officer_rows(num_other_officers: int):
        #         other_officer_rows = []
        #         for i in range(num_other_officers):
        #             other_officer_rows.append(gr.Row(visible=True))
        #         for i in range(MAX_NUM_OTHER_OFFICERS - num_other_officers):
        #             other_officer_rows.append(gr.Row(visible=False))
        #         return other_officer_rows
        #     num_other_officers.change(fn=enable_other_officer_rows, inputs=num_other_officers, outputs=other_officer_rows)

        #     with gr.Row():
        #         gr.Markdown("Please review the entries before hitting the save button. The current version does not support any changes once these details are saved. Also, please ensure all entries are filled out to have the correct forms generated.")

        #     with gr.Row():
        #         save_role_btn = gr.Button(value="Save", variant="primary")

        #     with gr.Row():
        #         with gr.Column():
        #             save_role_message = gr.Textbox(label="Message", interactive=False)
        #             save_role_btn.click(
        #                 fn=save_role_data,
        #                 inputs=[
        #                     company_id, founder_ids, founder_names, director_ids, director_names, 
        #                     chairman_id, chairman_name, ceo_id, ceo_name, ceo_open_account, ceo_draw_account, 
        #                     cfo_id, cfo_name, cfo_open_account, cfo_draw_account, secretary_id, secretary_name, 
        #                     secretary_open_account, secretary_draw_account] + role_elems,
        #                 outputs=role_ids + [save_role_message]
        #             )
        #         with gr.Column():
        #             next_btn = gr.Button(value="Next")
        #             next_btn.click(fn=next_tab, inputs=tab2_id, outputs=tabs)

        # ### STOCKHOLDER DETAILS

        # with gr.Tab(label="4. Add Stock Ownership Details", id=3):
        #     with gr.Row(visible=VISIBLE):
        #         tab3_id = gr.Number(label="Tab ID", interactive=False, value=3)

        #     with gr.Row():
        #         gr.Markdown("## Details of ownership by various individuals and corporations.")

        #     with gr.Row(variant="panel"):
        #         with gr.Column():
        #             purchase_price = gr.Number(label="Stock Purchase Price", value=0.00001, step=0.00001, minimum=0)            
        #         with gr.Column():
        #             issue_date = Calendar(label="Stocks Issue Date", value=datetime.today(), type="datetime")

        #         retrieved_elems.extend([purchase_price, issue_date])

        #     with gr.Row():
        #         gr.Markdown("# Please include only the stockholder names who own at least one share.")

        #     stockholder_rows = []
        #     with gr.Row(variant="panel") as stockholder_row:
        #         with gr.Column(visible=VISIBLE):
        #             stockholder_id = gr.Number(label="Stockholder ID", interactive=False)
        #         with gr.Column():
        #             stockholder_name = gr.Dropdown(label="Stockholder", choices=[], allow_custom_value=True)
        #             save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=stockholder_name)
        #             entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=stockholder_name)
        #         with gr.Column():
        #             num_shares = gr.Number(label="Number Of Shares", minimum=0, step=1)
        #             total_price = gr.Number(label="Aggregate Total Price", minimum=0)

        #             def update_total_price(purchase_price: float, num_shares: int) -> gr.Number:
        #                 return gr.Number(value=purchase_price * num_shares)
        #             num_shares.change(fn=update_total_price, inputs=[purchase_price, num_shares], outputs=total_price)

        #             vesting_toggle = gr.Radio(label="Vesting?", choices=["Yes", "No"], value="No")
        #         with gr.Column(visible=False) as vesting_panel:
        #             total_vesting_at_cliff = gr.Number(label="Total Vesting At Cliff", info="In Number of Shares", minimum=0)
        #             cliff_anniversary = gr.Textbox(label="Cliff", info="Yearly/Quarterly/Monthly/Custom")
        #             monthly_vesting_after_cliff = gr.Number(label="Monthly Vesting After Cliff", info="In Number Of Shares", minimum=0)

        #             def enable_vesting(vesting_toggle: str):
        #                 return gr.Column(visible=True) if vesting_toggle == "Yes" else gr.Column(visible=False)
        #             vesting_toggle.change(fn=enable_vesting, inputs=vesting_toggle, outputs=vesting_panel)

        #     stockholder_rows.append(stockholder_row)
        #     stockholder_ids.append(stockholder_id)
        #     stockholder_elems.append(stockholder_id)
        #     stockholder_elems.append(stockholder_name)
        #     stockholder_elems.append(num_shares)
        #     stockholder_elems.append(total_price)
        #     stockholder_elems.append(total_vesting_at_cliff)
        #     stockholder_elems.append(cliff_anniversary)
        #     stockholder_elems.append(monthly_vesting_after_cliff)

        #     retrieved_elems.append(vesting_toggle)
        #     retrieved_elems.extend(stockholder_elems)

        #     for i in range(MAX_NUM_ENTITIES-1):
        #         with gr.Row(variant="panel") as stockholder_row:
        #             with gr.Column(visible=VISIBLE):
        #                 _stockholder_id = gr.Number(label="Stockholder ID", interactive=False)
        #             with gr.Column():
        #                 _stockholder_name = gr.Dropdown(label="Stockholder", choices=[], allow_custom_value=True)
        #                 save_entity_btn.click(fn=update_dropdown, inputs=entity_id, outputs=_stockholder_name)
        #                 entity_id.change(fn=update_dropdown, inputs=entity_id, outputs=_stockholder_name)
        #             with gr.Column():
        #                 _num_shares = gr.Number(label="Number Of Shares", minimum=0, step=1)
        #                 _total_price = gr.Number(label="Aggregate Total Price", minimum=0)
        #                 _num_shares.change(fn=update_total_price, inputs=[purchase_price, _num_shares], outputs=_total_price)
        #                 _vesting_toggle = gr.Radio(label="Vesting?", choices=["Yes", "No"], value="No")
        #             with gr.Column(visible=False) as _vesting_panel:
        #                 _total_vesting_at_cliff = gr.Number(label="Total Vesting At Cliff", info="In Number of Shares", minimum=0)
        #                 _cliff_anniversary = gr.Textbox(label="Cliff", info="Yearly/Quarterly/Monthly/Custom")
        #                 _monthly_vesting_after_cliff = gr.Number(label="Monthly Vesting After Cliff", info="In Number Of Shares", minimum=0)
        #                 _vesting_toggle.change(fn=enable_vesting, inputs=_vesting_toggle, outputs=_vesting_panel)

        #         stockholder_rows.append(stockholder_row)
        #         stockholder_ids.append(_stockholder_id)
        #         stockholder_elems.append(_stockholder_id)
        #         stockholder_elems.append(_stockholder_name)
        #         stockholder_elems.append(_num_shares)
        #         stockholder_elems.append(_total_price)
        #         stockholder_elems.append(_total_vesting_at_cliff)
        #         stockholder_elems.append(_cliff_anniversary)
        #         stockholder_elems.append(_monthly_vesting_after_cliff)

        #         retrieved_elems.append(_vesting_toggle)
        #         retrieved_elems.extend(stockholder_elems[-7:])
            
        #     num_entities.change(fn=enable_entity_rows, inputs=num_entities, outputs=stockholder_rows)

        #     with gr.Row():
        #         gr.Markdown("Please review the entries before hitting the save button. The current version does not support any changes once these details are saved. Also, please ensure all entries are filled out to have the correct forms generated.")

        #     with gr.Row():
        #         save_stockholder_btn = gr.Button(value="Save", variant="primary")

        #     with gr.Row():
        #         with gr.Column():
        #             save_stockholder_message = gr.Textbox(label="Message", interactive=False)
        #             save_stockholder_btn.click(
        #                 fn=save_stockholder_data, 
        #                 inputs=[company_id, purchase_price, issue_date] + stockholder_elems, 
        #                 outputs=stockholder_ids + [save_stockholder_message]
        #             )
        #         with gr.Column():
        #             next_btn = gr.Button(value="Next")
        #             next_btn.click(fn=next_tab, inputs=tab3_id, outputs=tabs)

        #     def select_from_existing(existing_dropdown: str):
        #         company_id = int(existing_dropdown.split()[0])
        #         return gr.Number(value=company_id)
        #     existing_dropdown.change(fn=select_from_existing, inputs=existing_dropdown, outputs=company_id)
        #     company_id.change(fn=retrieve_all, inputs=company_id, outputs=retrieved_elems)

        ### SS4

        with gr.Tab(label="5. Form SS4", id=4):
            with gr.Row(visible=VISIBLE):
                tab4_id = gr.Number(
                    label="Tab ID",
                    interactive=False,
                    value=4
                )
            with gr.Row(variant="panel"):
                ss4_1 = gr.Textbox(label="1. Legal name of entity (or individual) for whom the EIN is being requested")
            with gr.Row(variant="panel"):
                with gr.Column():
                    ss4_2 = gr.Textbox(label="2. Trade name of business (if different from name on line 1)")
                with gr.Column():
                    ss4_3 = gr.Textbox(label="3. Executor, administrator, trustee, “care of” name")
            with gr.Row(variant="panel"):
                with gr.Column():
                    ss4_4a = gr.Textbox(label="4a. Mailing address (room, apt., suite no. and street, or P.O. box)")
                with gr.Column():
                    ss4_5a = gr.Textbox(label="5a. Street address (if different) (Don’t enter a P.O. box.)")
            with gr.Row(variant="panel"):
                with gr.Column():
                    ss4_4b = gr.Textbox(label="4b. City, state, and ZIP code (if foreign, see instructions)")
                with gr.Column():
                    ss4_5b = gr.Textbox(label="5b. City, state, and ZIP code (if foreign, see instructions)")
            with gr.Row(variant="panel"):
                ss4_6 = gr.Textbox(label="6. County and state where principal business is located")
            with gr.Row(variant="panel"):
                ss4_7a = gr.Textbox(label="7a. Name of responsible party")
                ss4_7b = gr.Textbox(label="7b. SSN, ITIN, or EIN")
            with gr.Row(variant="panel"):
                ss4_8a = gr.Radio(label="8a. Is this application for a limited liability company (LLC) (or a foreign equivalent)?", choices=["Yes", "No"])
                ss4_8b = gr.Textbox(label="8b. If  8a  is  “Yes,”  enter  the  number  of LLC members")
            with gr.Row(variant="panel"):
                ss4_8c = gr.Radio(label="8c. If 8a is “Yes,” was the LLC organized in the United States?", choices=["Yes", "No"])
            with gr.Row(variant="panel"):
                gr.Markdown("9a. Type of entity (check only one box). Caution: If 8a is “Yes,” see the instructions for the correct box to check")
                with gr.Column():
                    with gr.Row():
                        ss4_9aSoleProprietor = gr.Checkbox(label="Sole Proprietor (SSN)")
                        ss4_9aSoleProprietorINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aPartnership = gr.Checkbox(label="Partnership")
                    with gr.Row():
                        ss4_9aCorporation = gr.Checkbox(label="Corporation (enter form number to be filed)")
                        ss4_9aCorporationINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aPersonalServiceCorporation = gr.Checkbox(label="Personal service corporation")
                    with gr.Row():
                        ss4_9aChurch = gr.Checkbox(label="Church or church-controlled organization")
                    with gr.Row():
                        ss4_9aOtherNonprofitOrganization = gr.Checkbox(label="Other nonprofit organization (specify)")
                        ss4_9aOtherNonprofitOrganizationINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aOther = gr.Checkbox(label="Other (specify)")
                        ss4_9aOtherINFO = gr.Textbox(show_label=False)
                with gr.Column():
                    with gr.Row():
                        ss4_9aEstate = gr.Checkbox(label="Estate (SSN of decedent)")
                        ss4_9aEstateINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aPlanAdministrator = gr.Checkbox(label="Plan administrator (TIN)")
                        ss4_9aPlanAdministratorINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aTrust = gr.Checkbox(label="Trust (TIN of grantor)")
                        ss4_9aTrustINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_9aMilitary = gr.Checkbox(label="Military/National Guard")
                        ss4_9aStateGovernment = gr.Checkbox(label="State/local government")
                    with gr.Row():
                        ss4_9aFarmers = gr.Checkbox(label="Farmers’ cooperative")
                        ss4_9aFederalGovernment = gr.Checkbox(label="Federal government")
                    with gr.Row():
                        ss4_9aREMIC = gr.Checkbox(label="REMIC")
                        ss4_9aIndianTribalGovernments = gr.Checkbox(label="Indian tribal governments/enterprises")
                    with gr.Row():
                        ss4_9aGEN = gr.Textbox(label="Group Exemption Number (GEN) if any")
            with gr.Row(variant="panel"):
                with gr.Column():
                    gr.Markdown("9b. If a corporation, name the state or foreign country (if applicable) where incorporated")
                with gr.Column():
                    ss4_9bState = gr.Textbox(label="State")
                with gr.Column():
                    ss4_9bForeign = gr.Textbox(label="Foreign country")
            with gr.Row(variant="panel"):
                with gr.Column():
                    with gr.Row():
                        gr.Markdown("10. Reason for applying (check only one box)")
                    with gr.Row():
                        ss4_10StartNewBusiness = gr.Checkbox(label="Started new business (specify type)")
                        ss4_10StartNewBusinessINFO = gr.Textbox(show_label=False)
                    with gr.Row(visible=False):
                        ss4_10StartNewBusinessINFO2 = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_10HiredEmployees = gr.Checkbox(label="Hired employees (Check the box and see line 13.)")
                    with gr.Row():
                        ss4_10ComplianceIRS = gr.Checkbox(label="Compliance with IRS withholding regulations")
                    with gr.Row():
                        ss4_10Others = gr.Checkbox(label="Other (specify)")
                        ss4_10OthersINFO = gr.Textbox(show_label=False)
                with gr.Column():
                    with gr.Row():
                        ss4_10Banking = gr.Checkbox(label="Banking purpose (specify purpose)")
                        ss4_10BankingINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_10ChangedOrganization = gr.Checkbox(label="Changed type of organization (specify new type)")
                        ss4_10ChangedOrganizationINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_10PurchasedBusiness = gr.Checkbox(label="Purchased going business")
                    with gr.Row():
                        ss4_10CreatedTrust = gr.Checkbox(label="Created a trust (specify type)")
                        ss4_10CreatedTrustINFO = gr.Textbox(show_label=False)
                    with gr.Row():
                        ss4_10PensionPlan = gr.Checkbox(label="Created a pension plan (specify type)")
                        ss4_10PensionPlanINFO = gr.Textbox(show_label=False)
            with gr.Row(variant="panel"):
                with gr.Column():
                    ss4_11 = gr.Textbox(label="11. Date business started or acquired (month, day, year). See instructions.")
                with gr.Column():
                    ss4_12 = gr.Textbox(label="12. Closing month of accounting year")
            with gr.Row(variant="panel"):
                with gr.Column():
                    with gr.Row():
                        gr.Markdown("13. Highest number of employees expected in the next 12 months (enter -0- if none). If no employees expected, skip line 14")
                    with gr.Row():
                        ss4_13Agricultural = gr.Textbox(label="Agricultural")
                        ss4_13Household = gr.Textbox(label="Household")
                        ss4_13Other = gr.Textbox(label="Other")
                with gr.Column():
                    ss4_14 = gr.Checkbox(label="14. If you expect your employment tax liability to be $1,000 or less in a full calendar year and want to file Form 944 annually instead of Forms 941 quarterly, check here. (Your employment tax liability will generally be $1,000 or less if you expect to pay $5,000 or less, $6,536 or less if you’re in a U.S. territory, in total wages.) If you don’t check this box, you must file Form 941 for every quarter.")
            with gr.Row(variant="panel"):
                ss4_15 = gr.Textbox(label="15. First  date  wages  or  annuities  were  paid  (month,  day,  year).  Note:  If  applicant  is  a  withholding  agent,  enter  date  income  will  first  be  paid  to nonresident alien (month, day, year)")
            with gr.Row(variant="panel"):
                with gr.Column():
                    with gr.Row():
                        gr.Markdown("16. Check one box that best describes the principal activity of your business.")
                    with gr.Row():
                        ss4_16Construction = gr.Checkbox(label="Construction")
                        ss4_16Rental = gr.Checkbox(label="Rental & leasing")
                        ss4_16Transporting = gr.Checkbox(label="Transportation & warehousing")
                    with gr.Row():
                        ss4_16RealEstate = gr.Checkbox(label="Real estate")
                        ss4_16Manufacturing = gr.Checkbox(label="Manufacturing")
                        ss4_16Finance = gr.Checkbox(label="Finance & insurance")
                with gr.Column():
                    with gr.Row():
                        ss4_16HealthCare = gr.Checkbox(label="Health care & social assistance")
                        ss4_16WholesaleAgent = gr.Checkbox(label="Wholesale—agent/broker")
                    with gr.Row():
                        ss4_16Accommodation = gr.Checkbox(label="Accommodation & food service")
                        ss4_16WholesaleOther = gr.Checkbox(label="Wholesale—other")
                        ss4_16Retail = gr.Checkbox(label="Retail")
                    with gr.Row():
                        ss4_16Other = gr.Checkbox(label="Other (specify)")
                        ss4_16OtherINFO = gr.Textbox(show_label=False)
            with gr.Row(variant="panel"):
                ss4_17 = gr.Textbox(label="17. Indicate principal line of merchandise sold, specific construction work done, products produced, or services provided.")
            with gr.Row(variant="panel"):
                ss4_18 = gr.Radio(label="18. Has the applicant entity shown on line 1 ever applied for and received an EIN?", choices=["Yes", "No"])
                ss4_18EIN = gr.Textbox(label="If “Yes,” write previous EIN here")
            with gr.Row(variant="panel"):
                gr.Markdown("Third Party Designee")
                with gr.Column():
                    with gr.Row():
                        ss4_DesigneeName = gr.Textbox(label="Designee’s name")
                    with gr.Row():
                        ss4_DesigneeAddress = gr.Textbox(label="Address and ZIP code")
                with gr.Column():
                    with gr.Row():
                        ss4_DesigneeTelephone = gr.Textbox(label="Designee’s telephone number (include area code)")
                    with gr.Row():
                        ss4_DesigneeFax = gr.Textbox(label="Designee’s fax number (include area code)")
            with gr.Row(variant="panel"):
                with gr.Column():
                    with gr.Row():
                        ss4_ApplicantName = gr.Textbox(label="Applicant's Name and title")
                with gr.Column():
                    with gr.Row():
                        ss4_ApplicantTelephone = gr.Textbox(label="Applicant’s telephone number (include area code)")
                    with gr.Row():
                        ss4_ApplicantFax = gr.Textbox(label="Applicant’s fax number (include area code)")

            ss4_elems = [
                ss4_1, ss4_2, ss4_3, ss4_4a, ss4_4b, ss4_5a, ss4_5b, ss4_6, ss4_7a, ss4_7b, ss4_8a, ss4_8b, ss4_8c,
                ss4_9aSoleProprietor, ss4_9aSoleProprietorINFO, ss4_9aEstate, ss4_9aEstateINFO, ss4_9aPartnership,
                ss4_9aPlanAdministrator, ss4_9aPlanAdministratorINFO, ss4_9aCorporation, ss4_9aCorporationINFO,
                ss4_9aTrust, ss4_9aTrustINFO, ss4_9aPersonalServiceCorporation, ss4_9aMilitary, ss4_9aStateGovernment,
                ss4_9aChurch, ss4_9aFarmers, ss4_9aFederalGovernment, ss4_9aOtherNonprofitOrganization, 
                ss4_9aOtherNonprofitOrganizationINFO, ss4_9aREMIC, ss4_9aIndianTribalGovernments, ss4_9aOther,
                ss4_9aOtherINFO, ss4_9aGEN, ss4_9bState, ss4_9bForeign, ss4_10Banking, ss4_10BankingINFO, 
                ss4_10StartNewBusiness, ss4_10StartNewBusinessINFO, ss4_10ChangedOrganization, ss4_10ChangedOrganizationINFO,
                ss4_10PurchasedBusiness, ss4_10HiredEmployees, ss4_10CreatedTrust, ss4_10CreatedTrustINFO, 
                ss4_10ComplianceIRS, ss4_10PensionPlan, ss4_10PensionPlanINFO, ss4_10Others, ss4_10OthersINFO, 
                ss4_11, ss4_12, ss4_13Agricultural, ss4_13Household, ss4_13Other, ss4_14, ss4_15, ss4_16HealthCare,
                ss4_16WholesaleAgent, ss4_16Construction, ss4_16Rental, ss4_16Transporting, ss4_16Accommodation, 
                ss4_16WholesaleOther, ss4_16Retail, ss4_16RealEstate, ss4_16Manufacturing, ss4_16Finance, 
                ss4_16Other, ss4_16OtherINFO, ss4_17, ss4_18, ss4_18EIN, ss4_DesigneeName, ss4_DesigneeTelephone, 
                ss4_DesigneeAddress, ss4_DesigneeFax, ss4_ApplicantName, ss4_ApplicantTelephone, ss4_ApplicantFax
            ]

            with gr.Row():
                with gr.Column():
                    create_ss4_btn = gr.Button(value="Create SS4", variant="primary")
                with gr.Column():
                    ss4_output = gr.File(show_label=False)
                    create_ss4_btn.click(fn=fill_ss4_1, inputs=ss4_elems, outputs=ss4_output)

            with gr.Row():
                next_btn = gr.Button(value="Next")
                # next_btn.click(fn=next_tab, inputs=tab4_id, outputs=tabs)

        ### DOWNLOADS

        # with gr.Tab(label="6. Generate Outputs", id=5):
        #     with gr.Row(visible=VISIBLE):
        #         tab5_id = gr.Number(label="Tab ID", interactive=False, value=5)

        #     with gr.Row():
        #         gr.Markdown("## Download current and pre-existing forms.")

        #     with gr.Row(variant="panel"):
        #         download_checkbox = gr.CheckboxGroup(choices=forms, show_label=False, info="Select Forms to Create",)

        #     with gr.Row():
        #         submit_btn = gr.Button(value="Submit", variant="primary")

        #     with gr.Row(variant="panel"):
        #         file_explorer = gr.FileExplorer(root_dir="./outputs/", glob="**/*.*", label="Select the folders below to select files to view.")
        #         files = gr.Files(file_types=[".docx", ".pdf", ".zip"], label="Each file corresponds to the folder(s) selected.", visible=False)

        #         def update_files(file_explorer):
        #             if file_explorer:
        #                 return gr.Files(visible=True, value=file_explorer)
        #             return gr.Files(visible=False)
        #         file_explorer.change(fn=update_files, inputs=file_explorer, outputs=files)
            
        #     submit_btn.click(
        #         fn=update_document,
        #         inputs=[company_id, download_checkbox],
        #         outputs=file_explorer
        #     ).then(
        #         fn=update_root_dir,
        #         outputs=file_explorer
        #     )

        # ### Invoice

        # with gr.Tab(label="Invoice Estimate"):
        #     gr.Markdown("# Invoice Estimate")
        #     gr.Markdown("## This page allows a quick estimate of invoice, pricing. All numbers are mentioned in USD.")
            
        #     gr.Markdown("#### Base Price for Service  (USD) ")
        #     base_price = gr.Number(label="Review & update the base price as required", value=1900)

        #     gr.Markdown("### Agent Requirement")
        #     with gr.Row():
        #         states_input_agent = gr.Dropdown(label="Select states with Agent Requirement ", choices=states, multiselect=True)

        #     agent_state_rows = []
        #     agent_state_names = []
        #     agent_state_values = []
        #     with gr.Accordion(label="Expand for prices", open=False):
        #         for i in range(len(states)):
        #             with gr.Row(visible=False) as agent_state_row:
        #                 state_name = gr.Textbox(label="State", visible=False)
        #                 state_value = gr.Number(label="Price", visible=False)
        #                 agent_state_names.append(state_name)
        #                 agent_state_values.append(state_value)
        #             agent_state_rows.append(agent_state_row)
                
        #         def enable_agent_state_rows(states_input_agent: List[str]):
        #             agent_state_rows = []
        #             agent_state_names = []
        #             agent_state_values = []
        #             for state in states_input_agent:
        #                 agent_state_rows.append(gr.Row(visible=True))
        #                 agent_state_names.append(gr.Textbox(value=state, visible=True))
        #                 agent_state_values.append(gr.Number(value=99, visible=True))
        #             while len(agent_state_rows) < len(states):
        #                 agent_state_rows.append(gr.Row(visible=False))
        #             while len(agent_state_names) < len(states):
        #                 agent_state_names.append(
        #                     gr.Textbox(value=None, visible=False))
        #             while len(agent_state_values) < len(states):
        #                 agent_state_values.append(
        #                     gr.Number(value=0, visible=False))
        #             return agent_state_rows + agent_state_names + agent_state_values
        #         states_input_agent.change(fn=enable_agent_state_rows, inputs=states_input_agent, outputs=agent_state_rows + agent_state_names + agent_state_values)

        #     gr.Markdown("#### States with Registeration Requirement ")
        #     with gr.Row():    
        #         states_input_regi = gr.Dropdown(label="Select states with Registration Requirement", choices=states, multiselect=True)

        #     registration_state_rows = []
        #     registration_state_names = []
        #     registration_state_values = []
        #     with gr.Accordion(label="Expand for prices", open=False):
        #         for i in range(len(states)):
        #             with gr.Row(visible=False) as registration_state_row:
        #                 state_name = gr.Textbox(label="State", visible=False)
        #                 state_value = gr.Number(label="Price", visible=False)
        #                 registration_state_names.append(state_name)
        #                 registration_state_values.append(state_value)
        #             registration_state_rows.append(registration_state_row)
                
        #         def enable_registration_state_rows(states_input_regi: List[str]):
        #             registration_state_rows = []
        #             registration_state_names = []
        #             registration_state_values = []
        #             for state in states_input_regi:
        #                 registration_state_rows.append(gr.Row(visible=True))
        #                 registration_state_names.append(gr.Textbox(value=state, visible=True))
        #                 registration_state_values.append(gr.Number(value=99, visible=True))
        #             while len(registration_state_rows) < len(states):
        #                 registration_state_rows.append(gr.Row(visible=False))
        #             while len(registration_state_names) < len(states):
        #                 registration_state_names.append(gr.Textbox(value=None, visible=False))
        #             while len(registration_state_values) < len(states):
        #                 registration_state_values.append(gr.Number(value=0, visible=False))
        #             return registration_state_rows + registration_state_names + registration_state_values
        #         states_input_regi.change(fn=enable_registration_state_rows, inputs=states_input_regi, outputs=registration_state_rows + registration_state_names + registration_state_values)

        #     gr.Markdown("#### Requirement for US Address ")
        #     with gr.Row():
        #         us_address = gr.Radio(choices=["Yes", "No"], label="Select Yes, if you need a US Address. We help with providing a mail stop for up to 6 months.")
        #         us_address_value = gr.Number(label="The default value for this mail stop is $200. The price can be updated below.", value=200)

        #     gr.Markdown("#### Stock Option Plan Reserve ")
        #     with gr.Row():
        #         sop_reserve = gr.Radio(choices=["Yes", "No"], label="Would you like an SOP reserve?  Select the applicable option. ")
        #         sop_reserve_value = gr.Number(label="The default price (USD) for this is $1000, if you'd like to update it, you can review and update below.", value=1000)

        #     gr.Markdown("#### Founders")
        #     with gr.Row():
        #         add_founder = gr.Number(label="What is the total number of founders?", value=1)
        #         add_founder_value = gr.Number(label="The default package includes one founder, each additional founder costs extra. The price (USD) per additional founder is mentioned below and can be edited.", value=300)

        #     gr.Markdown("#### Document Processing Preference ")
        #     with gr.Row():
        #         docsign = gr.Radio(choices=["Yes", "No"], label="Would you prefer processing the documents via Docusign? Please review the options below. ", value="Yes")
        #         docsign_value = gr.Number(label="Please update the price (USD) if the charge is different from default value. ", value=50)

        #     with gr.Row():
        #         generate_button = gr.Button("Generate Invoice")

        #     with gr.Row():
        #         outputs = gr.File(show_label=False)
            
        #         inputs = [
        #             base_price, 
        #             states_input_agent, 
        #             states_input_regi, 
        #             us_address, 
        #             us_address_value, 
        #             sop_reserve, 
        #             sop_reserve_value, 
        #             add_founder, 
        #             add_founder_value, 
        #             docsign, 
        #             docsign_value
        #         ] + agent_state_values + registration_state_values
                
        #         generate_button.click(
        #             fn=create_invoice, 
        #             inputs=inputs, 
        #             outputs=outputs
        #         )


demo.queue().launch(
    # root_path="/questionnaire", 
    # server_port=7860, 
    # show_error=True, 
    # show_api=False,
    # ssl_certfile="cert.pem",
    # ssl_keyfile="key.pem",
    # ssl_verify=False,
    # auth=("user1", "decoy_tarcaz_1234")
)
